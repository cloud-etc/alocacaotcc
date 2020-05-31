import datetime
from django.contrib.admin.models import LogEntry
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.generic import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

from alocar.forms.alocar.forms import AddAlocForm
from alocar.models.alocar.models import Alocar
from alocar.models.sala.models import Sala
from alocar.models.turma.models import Turma

import logging

logger = logging.getLogger('django')

def comum(request):
    logEntry = LogEntry.objects.all()
    context = {
        "logEntry": logEntry
    }
    return render(request, "alocar/comum.html", context)

"""
funcao que da acessoa pagin a principal do sistema
e ao dashboard
"""
def home(request):
    context = { }
    return render(request, 'alocar/home.html', context)


"""
funcao para cadastrar alocacao
"""
@login_required
def listSalaTurma(request, *args, **kwargs):
    """
    verifica se o usuario tem perimissao, para fazer a operacao
    """
    varios = request.GET.get('varios', None)

    if varios:
        turmapage = Turma.objects.filter(turma__icontains=varios) | \
                    Turma.objects.filter(curso__icontains=varios) | \
                    Turma.objects.filter(periodo__icontains=varios) | \
                    Turma.objects.filter(disciplina__icontains=varios) | \
                    Turma.objects.filter(professor__icontains=varios)
    else:
        # turmapage = Turma.objects.all()
        turmapage = Turma.objects.filter(alocada=False)


    varios1 = request.GET.get('varios1', None)

    if varios1:
        salapage =  Sala.objects.filter(bloco__bloco__icontains=varios1) | \
                    Sala.objects.filter(sala__icontains=varios1)
    else:
        salapage =  Sala.objects.filter(

            disponivel=True
        )
        

    alocar = Alocar.objects.all()

    context = {}
    template_name = 'alocar/listsalaturma.html'

    form = AddAlocForm(request.POST or None)

    if form.is_valid():
        
        turma = form.cleaned_data['turma']
        sala = form.cleaned_data['sala']
        dia = form.cleaned_data['dia']
        horario = form.cleaned_data['horario']

        alocacao = form.save(commit=False)
        alocacao.user = request.user


        if alocacao.maior():
            messages.info(
                request, 'Quantidade de alunos é maior do que a capacidade da sala')
            return redirect('alocar:listalocacao')


        if alocacao.turma_computador():
            if alocacao.sala_computador():
                messages.info(
                    request, 'Essa turma precisa de computador')
                return redirect('alocar:listalocacao')

        
        if alocacao.turma_internet():
            if alocacao.sala_internet():
                messages.info(
                request, 'Essa turma precisa de internet')
                return redirect('alocar:listalocacao')


        if alocacao.turma_projetor():
            if alocacao.sala_projetor():
                messages.info(
                    request, 'Essa turma precisa de projetor')
                return redirect('alocar:listalocacao')


        query = Alocar.objects.filter(dia=dia) & \
                Alocar.objects.filter(horario=horario) & \
                Alocar.objects.filter(sala=sala)
        if query:
            messages.info(request, 'Essa sala já está alocada para esse dia e horário')
            return redirect('alocar:listalocacao')


        query = Alocar.objects.filter(dia=dia) & \
                Alocar.objects.filter(horario=horario) & \
                Alocar.objects.filter(turma__professor=turma.professor)
        if query:
            messages.info(
                request, 'Esse Professor já está em uma turma para esse dia e horário')
            return redirect('alocar:listalocacao')


        query = Alocar.objects.filter(dia=dia) & \
            Alocar.objects.filter(horario=horario) & \
            Alocar.objects.filter(turma=turma)
        if query:
            messages.info(
                request, 'Essa turma já está alocada para esse dia e horário')
            return redirect('alocar:listalocacao')


        query = Alocar.objects.filter(turma=turma) & \
                Alocar.objects.filter(sala=sala)
        if query:
            messages.info(request, 'Essa turma já está alocada')
            return redirect('alocar:listalocacao')

        else:
            alocacao.turma.esta_alocada()
            alocacao.save()
            form = AddAlocForm()
            return redirect('alocar:listalocacao')



    paginator = Paginator(turmapage, 30)    
    page = request.GET.get('page')
    turmapage = paginator.get_page(page)

    paginator1 = Paginator(salapage, 30)
    page1 = request.GET.get('page')
    salapage = paginator1.get_page(page1)

    context['form'] = form
    context['turmapage'] = turmapage
    context['salapage'] = salapage
    return render(request, template_name, context)


"""
funcao para listar as alocacoes
"""


@login_required
def listAlocacao(request):
    """
    pesquisa no banco dados
    """
    logger.info('Acessaram a listagem de alocação')
    varios = request.GET.get('varios', None)
    context = {}
    if varios:
        alocacoes = Alocar.objects.filter(turma__curso__icontains=varios) | \
                    Alocar.objects.filter(turma__periodo__icontains=varios) | \
                    Alocar.objects.filter(turma__disciplina__icontains=varios) | \
                    Alocar.objects.filter(turma__professor__icontains=varios) | \
                    Alocar.objects.filter(dia__icontains=varios) | \
                    Alocar.objects.filter(horario__horario__icontains=varios) | \
                    Alocar.objects.filter(sala__sala__icontains=varios)
    else:
        alocacoes = Alocar.objects.all()

    paginator = Paginator(alocacoes, 30)

    page = request.GET.get('page')

    alocapage = paginator.get_page(page)

    context = {'alocacoes': alocacoes, 'alocapage': alocapage}
    return render(request, 'alocar/listalocacao.html', context)



"""
funcao para alterar alocacao
"""
@login_required
def altAlocacao(request, id):
    """
        verifica se o usuario tem perimissao, para fazer a operacao
    """
    if not request.user.has_perm('alocar.change_alocar'):
        return render(request, 'alocar/permissao5.html')

    alocacao = get_object_or_404(Alocar, pk=id)

    form = AddAlocForm(request.POST or None, request.FILES or None,
                       instance=alocacao)

    if form.is_valid():

        alocacao = form.save(commit=False)
        alocacao.turma = form.cleaned_data['turma']
        alocacao.sala = form.cleaned_data['sala']
        alocacao.dia = form.cleaned_data['dia']
        alocacao.horario = form.cleaned_data['horario']

        if alocacao.maior():
            messages.info(
                request, 'Quantidade de alunos é maior do que a capacidade da sala')
            return redirect('alocar:listalocacao')

        if alocacao.turma_computador():
            if alocacao.sala_computador():
                messages.info(
                    request, 'Essa turma precisa de computador')
                return redirect('alocar:listalocacao')

        if alocacao.turma_internet():
            if alocacao.sala_internet():
                messages.info(
                    request, 'Essa turma precisa de internet')
                return redirect('alocar:listalocacao')

        if alocacao.turma_projetor():
            if alocacao.sala_projetor():
                messages.info(
                    request, 'Essa turma precisa de projetor')
                return redirect('alocar:listalocacao')

        query = Alocar.objects.filter(dia=alocacao.dia) & \
                Alocar.objects.filter(sala=alocacao.sala) & \
                Alocar.objects.filter(horario=alocacao.horario)
        if query:
            messages.info(request, 'Essa sala já está alocada para esse dia e horário')
            return redirect('alocar:listalocacao')


        query = Alocar.objects.filter(dia=alocacao.dia) & \
                Alocar.objects.filter(horario=alocacao.horario) & \
                Alocar.objects.filter(turma=alocacao.turma)
        if query:
            messages.info(
                request, 'Essa turma já está alocada para esse dia e horário')
            return redirect('alocar:listalocacao')

        query = Alocar.objects.filter(turma=alocacao.turma) & \
                Alocar.objects.filter(sala=alocacao.sala)
        if query:
            messages.info(request, 'Essa turma já está alocada')
            return redirect('alocar:listalocacao')

        query = Alocar.objects.filter(turma=alocacao.turma)


        if query:
            alocacao.turma.esta_alocada()
            alocacao.save()
            return redirect('alocar:listalocacao')
        else:
            form.turma.esta_alocada()
            form.save()
            return redirect('alocar:listalocacao')

    return render(request, 'alocar/altalocacao.html', {'form': form})




"""
funcao para deletar alocacao
"""
@login_required
def delAlocacao(request, id):
    """
     verifica se o usuario tem perimissao, para fazer a operacao
    """
    if not request.user.has_perm('alocar.delete_alocar'):
        return render(request, 'alocar/permissao5.html')

    context = {}
    alocar = get_object_or_404(Alocar, pk=id)

    if request.user.is_superuser:
        if request.method == 'POST':
            alocar.turma.nao_alocada()
            alocar.delete()
            return redirect('alocar:listalocacao')

    else:        

        if alocar.user == request.user:
            if request.method == 'POST':
                alocar.turma.nao_alocada()
                alocar.delete()
                return redirect('alocar:listalocacao')

        else:
            if not alocar.user == request.user:
                messages.info(
                request, 'Você NÃO pode apagar essa alocação. Porque foi feita por outro usuário')
                return redirect('alocar:listalocacao')
            

    context['alocar'] = alocar
    return render(request, 'alocar/delalocacao.html')



"""
consultar turmas sem alocação
"""
@login_required
def consultarTurmas(request):
    turmas = Turma.objects.raw(
        'SELECT alocar_turma.id, alocar_turma.turma, alocar_turma.curso '
        'FROM alocar_alocar INNER JOIN alocar_turma ON alocar_turma.id <> alocar_alocar.turma_id')

    context = { }
    context['consultarturmas'] = turmas
    return render(request, 'alocar/consultarturmas.html', context)



def permissao1(request):
    context = { }
    return render(request, 'alocar/permissao1.html', context)


"""
funcao que EXPORTA TODAS AS ALOCAÇÕES
"""
class Export_alocacoes(TemplateView):
    def get(self, request, *args, **kwargs):
        query = Alocar.objects.all()
        wb = Workbook()
        bandera = True
        cont = 1
        controlador = 4

        for q in query:
            # if bandera:
            ws = wb.active
            ws.title = 'Alocações'
            # bandera = False
            # else:
            # ws = wb.create_sheet('Hoja')

            ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B1'].fill = PatternFill(start_color='66CCCC', end_color='66FFCC', fill_type="solid")
            ws['B1'].font = Font(name='Calibri', size=12, bold=True)
            ws['B1'] = 'ALOCAÇÃO DE TURMAS'

            ws.merge_cells('B1:J1')

            ws.row_dimensions[1].height = 25

            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 20

            ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['B3'].font = Font(name='Calibri', size=10, bold=True)
            ws['B3'] = 'TURMA'

            ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['C3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['C3'].font = Font(name='Calibri', size=10, bold=True)
            ws['C3'] = 'BLOCO'

            ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['D3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['D3'].font = Font(name='Calibri', size=10, bold=True)
            ws['D3'] = 'SALA'

            ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['E3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['E3'].font = Font(name='Calibri', size=10, bold=True)
            ws['E3'] = 'CURSO'

            ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['F3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['F3'].font = Font(name='Calibri', size=10, bold=True)
            ws['F3'] = 'PERIODO'

            ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['G3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['G3'].font = Font(name='Calibri', size=10, bold=True)
            ws['G3'] = 'DISCIPLINA'

            ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['H3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['H3'].font = Font(name='Calibri', size=10, bold=True)
            ws['H3'] = 'PROFESSOR'

            ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['I3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['I3'].font = Font(name='Calibri', size=10, bold=True)
            ws['I3'] = 'DIA'

            ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['J3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['J3'].font = Font(name='Calibri', size=10, bold=True)
            ws['J3'] = 'HORÁRIO'

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=2).value = q.turma.turma
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=3).value = q.sala.bloco.bloco
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=4).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=4).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=4).value = q.sala.sala
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=5).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=5).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=5).value = q.turma.curso
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=6).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=6).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=6).value = q.turma.periodo
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=7).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=7).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=7).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=7).value = q.turma.disciplina
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=8).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=8).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=8).value = q.turma.professor
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=9).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=9).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=9).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=9).value = q.dia
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=10).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=10).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=10).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=10).value = q.horario.horario
            # fim do bloco

            controlador += 1
            # cont += 1


        relatorio = "SAT - Alocações de Turmas.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(relatorio)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


export_alocacoes = Export_alocacoes.as_view()



"""
FUNÇÃO DE RELATÓRIO FILTRADO
"""

class Relatorios(TemplateView):
    def get(self, request, *args, **kwargs):
        varios = request.GET.get('filtrado')
        print(request.GET)
        query = Alocar.objects.filter(turma__curso__icontains=varios) | \
                    Alocar.objects.filter(turma__periodo__icontains=varios) | \
                    Alocar.objects.filter(turma__disciplina__icontains=varios) | \
                    Alocar.objects.filter(turma__professor__icontains=varios) | \
                    Alocar.objects.filter(dia__icontains=varios) | \
                    Alocar.objects.filter(horario__horario__icontains=varios) | \
                    Alocar.objects.filter(sala__sala__icontains=varios)

        wb = Workbook()
        bandera = True
        cont = 1
        controlador = 4

        for q in query:
            # if bandera:
            ws = wb.active
            ws.title = 'Alocações'
            # bandera = False
            # else:
            # ws = wb.create_sheet('Hoja')

            ws['B1'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B1'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B1'].fill = PatternFill(start_color='66CCCC', end_color='66FFCC', fill_type="solid")
            ws['B1'].font = Font(name='Calibri', size=12, bold=True)
            ws['B1'] = 'ALOCAÇÃO DE TURMAS'

            ws.merge_cells('B1:J1')

            ws.row_dimensions[1].height = 25

            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 25
            ws.column_dimensions['H'].width = 20
            ws.column_dimensions['I'].width = 10
            ws.column_dimensions['J'].width = 20

            ws['B3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['B3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['B3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['B3'].font = Font(name='Calibri', size=10, bold=True)
            ws['B3'] = 'TURMA'

            ws['C3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['C3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['C3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['C3'].font = Font(name='Calibri', size=10, bold=True)
            ws['C3'] = 'BLOCO'

            ws['D3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['D3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['D3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['D3'].font = Font(name='Calibri', size=10, bold=True)
            ws['D3'] = 'SALA'

            ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['E3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['E3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['E3'].font = Font(name='Calibri', size=10, bold=True)
            ws['E3'] = 'CURSO'

            ws['F3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['F3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['F3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['F3'].font = Font(name='Calibri', size=10, bold=True)
            ws['F3'] = 'PERIODO'

            ws['G3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['G3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['G3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['G3'].font = Font(name='Calibri', size=10, bold=True)
            ws['G3'] = 'DISCIPLINA'

            ws['H3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['H3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['H3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['H3'].font = Font(name='Calibri', size=10, bold=True)
            ws['H3'] = 'PROFESSOR'

            ws['I3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['I3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['I3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['I3'].font = Font(name='Calibri', size=10, bold=True)
            ws['I3'] = 'DIA'

            ws['J3'].alignment = Alignment(horizontal="center", vertical="center")
            ws['J3'].border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                                     top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            ws['J3'].fill = PatternFill(start_color='66CFCC', end_color='66CFCC', fill_type="solid")
            ws['J3'].font = Font(name='Calibri', size=10, bold=True)
            ws['J3'] = 'HORÁRIO'

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=2).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=2).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=2).value = q.turma.turma
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=3).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=3).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=3).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=3).value = q.sala.bloco.bloco
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=4).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=4).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=4).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=4).value = q.sala.sala
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=5).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=5).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=5).value = q.turma.curso
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=6).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=6).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=6).value = q.turma.periodo
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=7).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=7).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=7).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=7).value = q.turma.disciplina
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=8).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=8).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=8).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=8).value = q.turma.professor
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=9).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=9).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=9).font = Font(name='Calibri', size=10)
            ws.cell(row=controlador, column=9).value = q.dia
            # fim do bloco

            # esse bloco de codigo é de acordo com a qtd de coluna, cada coluna um bloco
            ws.cell(row=controlador, column=10).alignment = Alignment(horizontal="center")
            ws.cell(row=controlador, column=10).border = Border(left=Side(border_style="thin"),
                                                               right=Side(border_style="thin"),
                                                               top=Side(border_style="thin"),
                                                               bottom=Side(border_style="thin"))
            ws.cell(row=controlador, column=10).font = Font(name='Calibri', size=8)
            ws.cell(row=controlador, column=10).value = q.horario.horario
            # fim do bloco

            controlador += 1
            # cont += 1

        relatorio = "SAT - Alocações de Turmas.xlsx"
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename = {0}".format(relatorio)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response


relatorios = Relatorios.as_view()

@login_required
def telapararelatorio(request):
    varios = request.GET.get('varios', None)
    if varios:
        alocacoes = Alocar.objects.filter(turma__curso__icontains=varios) | \
                    Alocar.objects.filter(turma__periodo__icontains=varios) | \
                    Alocar.objects.filter(turma__disciplina__icontains=varios) | \
                    Alocar.objects.filter(turma__professor__icontains=varios) | \
                    Alocar.objects.filter(dia__icontains=varios) | \
                    Alocar.objects.filter(horario__horario__icontains=varios) | \
                    Alocar.objects.filter(sala__sala__icontains=varios)
    else:
        alocacoes = Alocar.objects.all()

    template_name = 'alocar/relatorios.html'

    paginator = Paginator(alocacoes, 30)

    page = request.GET.get('page')

    alocapage = paginator.get_page(page)

    context = {'alocacoes': alocacoes, 'alocapage': alocapage}
    return render(request, template_name, context)






















